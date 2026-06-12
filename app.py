"""
刀模管理系统 - 云端数据库版
后端 API (Flask + SQLite)
"""

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from datetime import datetime
import sqlite3
import os

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

DATABASE = 'blade_molds.db'

def get_db():
    """获取数据库连接"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """初始化数据库表"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 刀模主表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS blades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,        -- 编号
            customer_code TEXT,                -- 客户料号
            cut_codes TEXT,                    -- 裁切过的料号
            spec TEXT,                         -- 规格
            blade_type TEXT,                   -- 刀片类型
            size REAL,                        -- 尺寸
            angle TEXT,                        -- 角度
            hole_count TEXT,                   -- 穴数
            quantity INTEGER,                  -- 数量
            total_cuts INTEGER DEFAULT 0,      -- 累计裁切数
            status TEXT DEFAULT '空闲',        -- 状态：空闲/领用中/样品
            remark TEXT,                       -- 备注
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 领用记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS borrow_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            blade_id INTEGER NOT NULL,
            borrower TEXT NOT NULL,            -- 领用人
            purpose TEXT,                       -- 用途
            borrow_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            return_date TIMESTAMP,              -- 归还日期
            status TEXT DEFAULT '领用中',       -- 领用中/已归还
            FOREIGN KEY (blade_id) REFERENCES blades(id)
        )
    ''')
    
    # 操作日志表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS operation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            blade_id INTEGER,
            operation TEXT NOT NULL,            -- 操作类型
            operator TEXT,                       -- 操作人
            detail TEXT,                         -- 详情
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()

# ============ 刀模管理 API ============

@app.route('/api/blades', methods=['GET'])
def get_blades():
    """获取所有刀模列表"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 支持搜索和筛选
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    
    sql = 'SELECT * FROM blades WHERE 1=1'
    params = []
    
    if search:
        sql += ' AND (code LIKE ? OR customer_code LIKE ? OR remark LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    
    if status:
        sql += ' AND status = ?'
        params.append(status)
    
    sql += ' ORDER BY code'
    
    cursor.execute(sql, params)
    blades = [dict(row) for row in cursor.fetchall()]
    
    # 获取当前领用信息
    for blade in blades:
        cursor.execute('''
            SELECT * FROM borrow_records 
            WHERE blade_id = ? AND status = '领用中'
            ORDER BY borrow_date DESC LIMIT 1
        ''', (blade['id'],))
        borrow = cursor.fetchone()
        if borrow:
            blade['borrow_info'] = dict(borrow)
        else:
            blade['borrow_info'] = None
    
    conn.close()
    return jsonify(blades)

@app.route('/api/blades', methods=['POST'])
def create_blade():
    """新增刀模"""
    data = request.json
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            INSERT INTO blades (code, customer_code, cut_codes, spec, blade_type, 
                              size, angle, hole_count, quantity, total_cuts, status, remark)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['code'],
            data.get('customer_code', ''),
            data.get('cut_codes', ''),
            data.get('spec', ''),
            data.get('blade_type', ''),
            data.get('size', 0),
            data.get('angle', ''),
            data.get('hole_count', ''),
            data.get('quantity', 1),
            data.get('total_cuts', 0),
            data.get('status', '空闲'),
            data.get('remark', '')
        ))
        
        blade_id = cursor.lastrowid
        
        # 记录操作日志
        cursor.execute('''
            INSERT INTO operation_logs (blade_id, operation, operator, detail)
            VALUES (?, '新增', ?, ?)
        ''', (blade_id, data.get('operator', '系统'), f"新增刀模 {data['code']}"))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'id': blade_id})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/blades/<int:id>', methods=['PUT'])
def update_blade(id):
    """更新刀模"""
    data = request.json
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            UPDATE blades SET 
                code = ?, customer_code = ?, cut_codes = ?, spec = ?,
                blade_type = ?, size = ?, angle = ?, hole_count = ?,
                quantity = ?, total_cuts = ?, status = ?, remark = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data['code'], data.get('customer_code', ''), data.get('cut_codes', ''),
            data.get('spec', ''), data.get('blade_type', ''), data.get('size', 0),
            data.get('angle', ''), data.get('hole_count', ''), data.get('quantity', 1),
            data.get('total_cuts', 0), data.get('status', '空闲'), data.get('remark', ''),
            id
        ))
        
        # 记录操作日志
        cursor.execute('''
            INSERT INTO operation_logs (blade_id, operation, operator, detail)
            VALUES (?, '更新', ?, ?)
        ''', (id, data.get('operator', '系统'), f"更新刀模 {data['code']}"))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/blades/<int:id>', methods=['DELETE'])
def delete_blade(id):
    """删除刀模"""
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # 获取刀模信息用于日志
        cursor.execute('SELECT code FROM blades WHERE id = ?', (id,))
        blade = cursor.fetchone()
        code = blade['code'] if blade else '未知'
        
        cursor.execute('DELETE FROM blades WHERE id = ?', (id,))
        
        # 记录操作日志
        cursor.execute('''
            INSERT INTO operation_logs (blade_id, operation, operator, detail)
            VALUES (?, '删除', ?, ?)
        ''', (id, '系统', f"删除刀模 {code}"))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

# ============ 领用/归还 API ============

@app.route('/api/borrow', methods=['POST'])
def borrow_blade():
    """领用刀模"""
    data = request.json
    blade_id = data['blade_id']
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # 检查刀模状态
        cursor.execute('SELECT code, status FROM blades WHERE id = ?', (blade_id,))
        blade = cursor.fetchone()
        
        if blade['status'] != '空闲':
            return jsonify({'success': False, 'error': '刀模当前不可用'}), 400
        
        # 创建领用记录
        cursor.execute('''
            INSERT INTO borrow_records (blade_id, borrower, purpose)
            VALUES (?, ?, ?)
        ''', (blade_id, data['borrower'], data.get('purpose', '')))
        
        # 更新刀模状态
        cursor.execute('''
            UPDATE blades SET status = '领用中', updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (blade_id,))
        
        # 记录操作日志
        cursor.execute('''
            INSERT INTO operation_logs (blade_id, operation, operator, detail)
            VALUES (?, '领用', ?, ?)
        ''', (blade_id, data['borrower'], f"领用刀模 {blade['code']}"))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/return', methods=['POST'])
def return_blade():
    """归还刀模"""
    data = request.json
    blade_id = data['blade_id']
    add_cuts = data.get('add_cuts', 0)  # 本次使用增加裁切数
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # 获取刀模信息
        cursor.execute('SELECT code, total_cuts FROM blades WHERE id = ?', (blade_id,))
        blade = cursor.fetchone()
        
        # 更新领用记录
        cursor.execute('''
            UPDATE borrow_records 
            SET status = '已归还', return_date = CURRENT_TIMESTAMP
            WHERE blade_id = ? AND status = '领用中'
        ''', (blade_id,))
        
        # 更新刀模状态
        cursor.execute('''
            UPDATE blades 
            SET status = '空闲', 
                total_cuts = total_cuts + ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (add_cuts, blade_id))
        
        # 记录操作日志
        cursor.execute('''
            INSERT INTO operation_logs (blade_id, operation, operator, detail)
            VALUES (?, '归还', ?, ?)
        ''', (blade_id, data.get('operator', '系统'), 
              f"归还刀模 {blade['code']}，增加裁切数 {add_cuts}"))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

# ============ 记录查询 API ============

@app.route('/api/records', methods=['GET'])
def get_records():
    """获取领用/归还记录"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT br.*, b.code as blade_code, b.customer_code
        FROM borrow_records br
        LEFT JOIN blades b ON br.blade_id = b.id
        ORDER BY br.borrow_date DESC
        LIMIT 100
    ''')
    
    records = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify(records)

@app.route('/api/logs', methods=['GET'])
def get_logs():
    """获取操作日志"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT ol.*, b.code as blade_code
        FROM operation_logs ol
        LEFT JOIN blades b ON ol.blade_id = b.id
        ORDER BY ol.created_at DESC
        LIMIT 100
    ''')
    
    logs = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify(logs)

# ============ 数据统计 API ============

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """获取统计数据"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 总数
    cursor.execute('SELECT COUNT(*) as total FROM blades')
    total = cursor.fetchone()['total']
    
    # 空闲数
    cursor.execute("SELECT COUNT(*) as free FROM blades WHERE status = '空闲'")
    free = cursor.fetchone()['free']
    
    # 领用中
    cursor.execute("SELECT COUNT(*) as borrowed FROM blades WHERE status = '领用中'")
    borrowed = cursor.fetchone()['borrowed']
    
    # 样品
    cursor.execute("SELECT COUNT(*) as sample FROM blades WHERE status = '样品'")
    sample = cursor.fetchone()['sample']
    
    # 总裁切数
    cursor.execute('SELECT SUM(total_cuts) as total_cuts FROM blades')
    total_cuts = cursor.fetchone()['total_cuts'] or 0
    
    conn.close()
    
    return jsonify({
        'total': total,
        'free': free,
        'borrowed': borrowed,
        'sample': sample,
        'total_cuts': total_cuts
    })

# ============ 静态页面 ============

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/init-data', methods=['POST'])
def init_data():
    """初始化数据（从Excel导入）"""
    from openpyxl import load_workbook
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传文件'}), 400
    
    file = request.files['file']
    wb = load_workbook(file)
    ws = wb.active
    
    conn = get_db()
    cursor = conn.cursor()
    
    success_count = 0
    error_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        if not row[0]:  # 编号为空跳过
            continue
        
        try:
            cursor.execute('''
                INSERT OR IGNORE INTO blades 
                (code, customer_code, cut_codes, spec, blade_type, size, angle, 
                 hole_count, quantity, total_cuts, status, remark)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                str(row[0]) if row[0] else '',
                str(row[1]) if row[1] else '',
                str(row[2]) if row[2] else '',
                str(row[3]) if row[3] else '',
                str(row[4]) if row[4] else '',
                float(row[5]) if row[5] else 0,
                str(row[6]) if row[6] else '',
                str(row[7]) if row[7] else '',
                int(row[8]) if row[8] else 1,
                int(row[9]) if row[9] else 0,
                str(row[10]) if row[10] else '空闲',
                str(row[11]) if row[11] else ''
            ))
            if cursor.rowcount > 0:
                success_count += 1
            else:
                error_count += 1
        except Exception as e:
            error_count += 1
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True, 
        'imported': success_count,
        'skipped': error_count
    })

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
